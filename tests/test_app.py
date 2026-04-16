"""Tests for the Flask web application."""

import io
import json
import os
import types
from unittest.mock import MagicMock, patch

import pytest

from app import app, build_args_from_form, get_cached_keys, process_file, TEMP_DIR, DELIMITER


@pytest.fixture
def client():
    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client


def parse_ndjson(data):
    """Parse NDJSON response bytes into a list of dicts."""
    lines = data.decode().strip().splitlines()
    return [json.loads(line) for line in lines if line.strip()]


# ---------------------------------------------------------------------------
# build_args_from_form
# ---------------------------------------------------------------------------

class TestBuildArgsFromForm:
    def test_defaults(self):
        with app.test_request_context():
            args = build_args_from_form({})
            assert args.generate_kml is False
            assert args.ascii_output is False

    def test_provider_flags_off_by_default(self):
        with app.test_request_context():
            args = build_args_from_form({})
            assert args.use_ip_db is False
            assert args.use_dnsbl is False
            assert args.use_ipinfo is False
            assert args.use_dnsdumpster is False
            assert args.use_abuseipdb is False
            assert args.use_proxycheck is False

    def test_provider_flags_enabled(self):
        with app.test_request_context():
            form = {"use_dnsbl": "on", "use_ipinfo": "on"}
            args = build_args_from_form(form)
            assert args.use_dnsbl is True
            assert args.use_ipinfo is True
            assert args.use_ip_db is False

    def test_api_keys_stripped(self):
        with app.test_request_context():
            form = {
                "use_abuseipdb": "on",
                "abuseipdb_api": "  abc123  ",
                "use_proxycheck": "on",
                "proxycheck_api": "",
            }
            args = build_args_from_form(form)
            assert args.abuseipdb_api == "abc123"
            assert args.proxycheck_api is None


# ---------------------------------------------------------------------------
# process_file
# ---------------------------------------------------------------------------

class TestProcessFile:
    def _make_args(self, **overrides):
        defaults = dict(
            generate_kml=False,
            ascii_output=False,
            use_ip_db=False, ip_db_key=None,
            use_dnsbl=False,
            use_ipinfo=False,
            use_dnsdumpster=False, dnsdumpster_api=None,
            use_abuseipdb=False, abuseipdb_api=None,
            use_proxycheck=False, proxycheck_api=None,
        )
        defaults.update(overrides)
        return types.SimpleNamespace(**defaults)

    def test_no_providers_selected(self):
        args = self._make_args()
        output, error = process_file("8.8.8.8", args)
        assert output is None
        assert "No providers selected" in error

    def test_no_ips_found(self):
        args = self._make_args(use_dnsbl=True)
        output, error = process_file("no ips here at all", args)
        assert output is None
        assert "No IP addresses found" in error

    def test_dnsdumpster_without_ipinfo(self):
        args = self._make_args(use_dnsdumpster=True)
        output, error = process_file("8.8.8.8", args)
        assert output is None
        assert "DNSDumpster requires IPInfo" in error

    @patch("app.create_providers")
    def test_extracts_ips_from_arbitrary_text(self, mock_create):
        provider = MagicMock()
        provider.name = "mock"
        provider.enabled = False
        provider.initialize.side_effect = lambda args: setattr(provider, 'enabled', True) or True
        provider.get_headers.return_value = ["Score"]
        provider.enrich.return_value = ["42"]

        mock_create.return_value = [provider]

        args = self._make_args(use_dnsbl=True)
        content = "some log: connected from 8.8.8.8 and also 1.1.1.1 done"
        output, error = process_file(content, args)

        assert error is None
        assert len(output) == 3  # header + 2 IPs
        assert output[0] == "IP;Score"
        assert output[1] == "8.8.8.8;42"
        assert output[2] == "1.1.1.1;42"

    @patch("app.create_providers")
    def test_deduplicates_ips(self, mock_create):
        provider = MagicMock()
        provider.name = "mock"
        provider.enabled = False
        provider.initialize.side_effect = lambda args: setattr(provider, 'enabled', True) or True
        provider.get_headers.return_value = ["X"]
        provider.enrich.return_value = ["1"]

        mock_create.return_value = [provider]

        args = self._make_args(use_dnsbl=True)
        output, error = process_file("8.8.8.8 8.8.8.8 1.1.1.1 8.8.8.8", args)

        assert error is None
        assert len(output) == 3  # header + 2 unique IPs

    @patch("app.create_providers")
    def test_csv_input_appends_enriched_columns(self, mock_create):
        provider = MagicMock()
        provider.name = "mock"
        provider.enabled = False
        provider.initialize.side_effect = lambda args: setattr(provider, 'enabled', True) or True
        provider.get_headers.return_value = ["Score"]
        provider.enrich.return_value = ["42"]

        mock_create.return_value = [provider]

        args = self._make_args(use_dnsbl=True)
        csv_content = "host;ip;port\nserver1;8.8.8.8;443\nserver2;1.1.1.1;80\n"
        output, error = process_file(csv_content, args)

        assert error is None
        assert output[0] == "host;ip;port;Score"
        assert output[1] == "server1;8.8.8.8;443;42"
        assert output[2] == "server2;1.1.1.1;80;42"

    @patch("app.create_providers")
    def test_csv_with_comma_delimiter(self, mock_create):
        provider = MagicMock()
        provider.name = "mock"
        provider.enabled = False
        provider.initialize.side_effect = lambda args: setattr(provider, 'enabled', True) or True
        provider.get_headers.return_value = ["X"]
        provider.enrich.return_value = ["ok"]

        mock_create.return_value = [provider]

        args = self._make_args(use_dnsbl=True)
        csv_content = "ip,label\n8.8.8.8,dns\n1.1.1.1,cloudflare\n"
        output, error = process_file(csv_content, args)

        assert error is None
        assert output[0] == "ip;label;X"
        assert output[1] == "8.8.8.8;dns;ok"
        provider.enrich.assert_any_call("8.8.8.8", {})
        provider.enrich.assert_any_call("1.1.1.1", {})

    @patch("app.create_providers")
    def test_csv_extracts_ip_from_cell_with_surrounding_text(self, mock_create):
        """CSV mode should pull the IP out of a cell even when it's embedded in text."""
        provider = MagicMock()
        provider.name = "mock"
        provider.enabled = False
        provider.initialize.side_effect = lambda args: setattr(provider, 'enabled', True) or True
        provider.get_headers.return_value = ["Score"]
        provider.enrich.return_value = ["42"]

        mock_create.return_value = [provider]

        args = self._make_args(use_dnsbl=True)
        csv_content = "id;source\nevt1;connection from 8.8.8.8\nevt2;hit 1.1.1.1 port 443\n"
        output, error = process_file(csv_content, args)

        assert error is None
        # Original rows preserved; enriched column appended
        assert output[0] == "id;source;Score"
        assert output[1] == "evt1;connection from 8.8.8.8;42"
        assert output[2] == "evt2;hit 1.1.1.1 port 443;42"
        # Provider was called with the IP extracted from the cell, not the full text
        provider.enrich.assert_any_call("8.8.8.8", {})
        provider.enrich.assert_any_call("1.1.1.1", {})

    @patch("app.create_providers")
    def test_csv_scans_all_cells_for_ip(self, mock_create):
        """CSV mode should scan every cell in each row, not only the detected column."""
        provider = MagicMock()
        provider.name = "mock"
        provider.enabled = False
        provider.initialize.side_effect = lambda args: setattr(provider, 'enabled', True) or True
        provider.get_headers.return_value = ["X"]
        provider.enrich.return_value = ["ok"]

        mock_create.return_value = [provider]

        args = self._make_args(use_dnsbl=True)
        # Row 1 has the IP in column 2 (note); row 2 has the IP in column 0 (name)
        csv_content = "name;email;note\nalice;alice@example.com;hit from 8.8.8.8\n1.1.1.1;bob@example.com;nothing\n"
        output, error = process_file(csv_content, args)

        assert error is None
        provider.enrich.assert_any_call("8.8.8.8", {})
        provider.enrich.assert_any_call("1.1.1.1", {})

    @patch("app.create_providers")
    def test_csv_no_ip_anywhere_returns_error(self, mock_create):
        """CSV with no IP-looking content anywhere should error with 'no IPs found'."""
        provider = MagicMock()
        provider.name = "mock"
        provider.enabled = False
        provider.initialize.side_effect = lambda args: setattr(provider, 'enabled', True) or True
        provider.get_headers.return_value = ["X"]
        provider.enrich.return_value = ["1"]

        mock_create.return_value = [provider]

        args = self._make_args(use_dnsbl=True)
        csv_content = "name;note\nalice;hello\nbob;world\n"
        output, error = process_file(csv_content, args)

        assert output is None
        assert "No IP addresses found" in error

    @patch("app.create_providers")
    def test_provider_returns_none(self, mock_create):
        """When provider.enrich returns None, empty values should be used."""
        provider = MagicMock()
        provider.name = "mock"
        provider.enabled = False
        provider.initialize.side_effect = lambda args: setattr(provider, 'enabled', True) or True
        provider.get_headers.return_value = ["A", "B"]
        provider.enrich.return_value = None

        mock_create.return_value = [provider]

        args = self._make_args(use_dnsbl=True)
        output, error = process_file("8.8.8.8", args)

        assert error is None
        assert output[1] == "8.8.8.8;;"

    @patch("app.create_providers")
    def test_provider_init_returns_false(self, mock_create):
        provider = MagicMock()
        provider.name = "badprov"
        provider.enabled = True
        provider.initialize.return_value = False

        mock_create.return_value = [provider]

        args = self._make_args(use_dnsbl=True)
        output, error = process_file("8.8.8.8", args)

        assert output is None
        assert "badprov" in error

    @patch("app.create_providers")
    def test_provider_init_calls_sys_exit(self, mock_create):
        provider = MagicMock()
        provider.name = "exitprov"
        provider.initialize.side_effect = SystemExit(1)

        mock_create.return_value = [provider]

        args = self._make_args(use_dnsbl=True)
        output, error = process_file("8.8.8.8", args)

        assert output is None
        assert "exitprov" in error

    @patch("app.create_providers")
    def test_progress_callback_called(self, mock_create):
        provider = MagicMock()
        provider.name = "mock"
        provider.enabled = False
        provider.initialize.side_effect = lambda args: setattr(provider, 'enabled', True) or True
        provider.get_headers.return_value = ["X"]
        provider.enrich.return_value = ["1"]

        mock_create.return_value = [provider]

        events = []
        def callback(event, data):
            events.append((event, data))

        args = self._make_args(use_dnsbl=True)
        process_file("8.8.8.8 1.1.1.1", args, progress_callback=callback)

        status_events = [e for e in events if e[0] == "status"]
        progress_events = [e for e in events if e[0] == "progress"]
        assert len(status_events) >= 1
        assert len(progress_events) == 2
        assert progress_events[0][1] == {"current": 1, "total": 2}
        assert progress_events[1][1] == {"current": 2, "total": 2}

    @patch("app.create_providers")
    def test_progress_callback_sets_provider_callback(self, mock_create):
        provider = MagicMock()
        provider.name = "mock"
        provider.enabled = False
        provider.initialize.side_effect = lambda args: setattr(provider, 'enabled', True) or True
        provider.get_headers.return_value = ["X"]
        provider.enrich.return_value = ["1"]

        mock_create.return_value = [provider]

        cb = lambda event, data: None
        args = self._make_args(use_dnsbl=True)
        process_file("8.8.8.8", args, progress_callback=cb)

        assert provider.progress_callback is cb


# ---------------------------------------------------------------------------
# Flask routes
# ---------------------------------------------------------------------------

class TestIndexRoute:
    def test_get_index(self, client):
        resp = client.get("/enrich-ip/")
        assert resp.status_code == 200
        assert b"enrich-ip" in resp.data
        assert b"input_file" in resp.data

    def test_index_contains_provider_checkboxes(self, client):
        resp = client.get("/enrich-ip/")
        html = resp.data.decode()
        assert 'name="use_dnsbl"' in html
        assert 'name="use_ip_db"' in html
        assert 'name="use_ipinfo"' in html
        assert 'name="use_dnsdumpster"' in html
        assert 'name="use_abuseipdb"' in html
        assert 'name="use_proxycheck"' in html

    def test_index_has_progress_bar(self, client):
        resp = client.get("/enrich-ip/")
        html = resp.data.decode()
        assert 'id="progress-section"' in html
        assert 'id="progress-fill"' in html


class TestCachedKeys:
    def test_no_cache_files(self, tmp_path):
        with patch("app.Path.home", return_value=tmp_path):
            cached = get_cached_keys()
        assert len(cached) == 0

    def test_detects_cached_key_file(self, tmp_path):
        (tmp_path / ".enrichip-abuseipdb-key").write_text("secret")
        with patch("app.Path.home", return_value=tmp_path):
            cached = get_cached_keys()
        assert "abuseipdb" in cached

    def test_ignores_directory(self, tmp_path):
        (tmp_path / ".enrichip-abuseipdb-key").mkdir()
        with patch("app.Path.home", return_value=tmp_path):
            cached = get_cached_keys()
        assert "abuseipdb" not in cached

    def test_index_shows_cached_note(self, client):
        with patch("app.get_cached_keys", return_value={"abuseipdb", "proxycheck"}):
            resp = client.get("/enrich-ip/")
        html = resp.data.decode()
        # Cached providers should show "API key cached" instead of input
        assert 'name="abuseipdb_api"' not in html
        assert 'name="proxycheck_api"' not in html
        # Non-cached providers should still show input
        assert 'name="ip_db_key"' in html

    def test_index_shows_input_when_not_cached(self, client):
        with patch("app.get_cached_keys", return_value=set()):
            resp = client.get("/enrich-ip/")
        html = resp.data.decode()
        assert 'name="ip_db_key"' in html
        assert 'name="abuseipdb_api"' in html
        assert 'name="proxycheck_api"' in html


class TestEnrichRoute:
    def test_no_file_returns_error_event(self, client):
        resp = client.post("/enrich-ip/enrich", data={})
        assert resp.status_code == 200
        events = parse_ndjson(resp.data)
        assert events[0]["event"] == "error"
        assert "No file" in events[0]["message"]

    def test_empty_file_returns_error_event(self, client):
        data = {"input_file": (io.BytesIO(b""), "empty.txt")}
        resp = client.post("/enrich-ip/enrich", data=data, content_type="multipart/form-data")
        events = parse_ndjson(resp.data)
        assert events[-1]["event"] == "error"
        assert "empty" in events[-1]["message"].lower()

    def test_no_providers_returns_error_event(self, client):
        data = {
            "input_file": (io.BytesIO(b"8.8.8.8\n"), "test.txt"),
        }
        resp = client.post("/enrich-ip/enrich", data=data, content_type="multipart/form-data")
        events = parse_ndjson(resp.data)
        error_events = [e for e in events if e["event"] == "error"]
        assert len(error_events) == 1
        assert "No providers selected" in error_events[0]["message"]

    @patch("app.process_file")
    def test_successful_enrichment(self, mock_process, client):
        mock_process.return_value = (["IP;Score", "8.8.8.8;42"], None)

        data = {
            "input_file": (io.BytesIO(b"8.8.8.8\n"), "test.txt"),
            "use_dnsbl": "on",
        }
        resp = client.post("/enrich-ip/enrich", data=data, content_type="multipart/form-data")
        assert resp.status_code == 200
        events = parse_ndjson(resp.data)
        done_events = [e for e in events if e["event"] == "done"]
        assert len(done_events) == 1
        assert "filename" in done_events[0]
        assert done_events[0]["headers"] == ["IP", "Score"]
        assert done_events[0]["rows"] == [["8.8.8.8", "42"]]

    @patch("app.process_file")
    def test_processing_error_returns_error_event(self, mock_process, client):
        mock_process.return_value = (None, "Something broke")

        data = {
            "input_file": (io.BytesIO(b"8.8.8.8\n"), "test.txt"),
            "use_dnsbl": "on",
        }
        resp = client.post("/enrich-ip/enrich", data=data, content_type="multipart/form-data")
        events = parse_ndjson(resp.data)
        error_events = [e for e in events if e["event"] == "error"]
        assert len(error_events) == 1
        assert "Something broke" in error_events[0]["message"]

    @patch("app.process_file")
    def test_excel_upload_parsed_to_csv(self, mock_process, client, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["host", "ip"])
        ws.append(["server1", "8.8.8.8"])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        mock_process.return_value = (["host;ip;Score", "server1;8.8.8.8;42"], None)

        data = {
            "input_file": (io.BytesIO(path.read_bytes()), "test.xlsx"),
            "use_dnsbl": "on",
        }
        resp = client.post("/enrich-ip/enrich", data=data, content_type="multipart/form-data")
        assert resp.status_code == 200
        # process_file should have been called with CSV-text content derived from the xlsx
        call_args = mock_process.call_args
        content_arg = call_args[0][0]
        assert "host;ip" in content_arg
        assert "server1;8.8.8.8" in content_arg

    def test_response_is_ndjson(self, client):
        data = {"input_file": (io.BytesIO(b"8.8.8.8\n"), "test.txt")}
        resp = client.post("/enrich-ip/enrich", data=data, content_type="multipart/form-data")
        assert "application/x-ndjson" in resp.content_type

    @patch("app.process_file")
    def test_streaming_includes_progress_events(self, mock_process, client):
        """When process_file calls the callback, progress events appear in the stream."""
        def fake_process(content, args, progress_callback=None):
            if progress_callback:
                progress_callback("status", {"message": "Working..."})
                progress_callback("progress", {"current": 1, "total": 2})
                progress_callback("progress", {"current": 2, "total": 2})
            return (["IP;Score", "8.8.8.8;42", "1.1.1.1;10"], None)

        mock_process.side_effect = fake_process

        data = {
            "input_file": (io.BytesIO(b"8.8.8.8\n1.1.1.1\n"), "test.txt"),
            "use_dnsbl": "on",
        }
        resp = client.post("/enrich-ip/enrich", data=data, content_type="multipart/form-data")
        events = parse_ndjson(resp.data)

        event_types = [e["event"] for e in events]
        assert "status" in event_types
        assert "progress" in event_types
        assert "done" in event_types


class TestDownloadRoute:
    def test_download_existing_file(self, client):
        # Create a temp file to download
        filename = "testdownload.csv"
        path = os.path.join(TEMP_DIR, filename)
        with open(path, "w") as f:
            f.write("IP;Score\n8.8.8.8;0\n")

        resp = client.get(f"/enrich-ip/download/{filename}")
        assert resp.status_code == 200
        assert b"IP;Score" in resp.data
        assert resp.headers["Content-Disposition"] == "attachment; filename=enriched.csv"

        os.unlink(path)

    def test_download_missing_file_404(self, client):
        resp = client.get("/enrich-ip/download/nonexistent.csv")
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Utils
# ---------------------------------------------------------------------------

class TestUtils:
    def test_is_valid_public_ip(self):
        from utils import is_valid_public_ip
        assert is_valid_public_ip("8.8.8.8") is True
        assert is_valid_public_ip("1.1.1.1") is True
        assert is_valid_public_ip("192.168.1.1") is False
        assert is_valid_public_ip("10.0.0.1") is False
        assert is_valid_public_ip("127.0.0.1") is False
        assert is_valid_public_ip("not-an-ip") is False
        assert is_valid_public_ip("") is False

    def test_is_private_ip(self):
        from utils import is_private_ip
        assert is_private_ip("192.168.1.1") is True
        assert is_private_ip("10.0.0.1") is True
        assert is_private_ip("8.8.8.8") is False
        assert is_private_ip("garbage") is False

    def test_extract_ips_from_text_ipv4(self):
        from utils import extract_ips_from_text
        result = extract_ips_from_text("connect from 8.8.8.8 and 1.1.1.1")
        assert result == ["8.8.8.8", "1.1.1.1"]

    def test_extract_ips_deduplicates(self):
        from utils import extract_ips_from_text
        result = extract_ips_from_text("8.8.8.8 8.8.8.8 1.1.1.1")
        assert result == ["8.8.8.8", "1.1.1.1"]

    def test_extract_ips_preserves_order(self):
        from utils import extract_ips_from_text
        result = extract_ips_from_text("first: 1.2.3.4 then: 5.6.7.8")
        assert result == ["1.2.3.4", "5.6.7.8"]

    def test_extract_ips_empty_text(self):
        from utils import extract_ips_from_text
        assert extract_ips_from_text("no addresses here") == []

    def test_extract_ips_ipv6(self):
        from utils import extract_ips_from_text
        result = extract_ips_from_text("addr: 2001:db8::1 end")
        assert "2001:db8::1" in result

    def test_extract_ips_mixed_content(self):
        from utils import extract_ips_from_text
        text = "Jan 01 log[123]: src=8.8.8.8 dst=1.1.1.1 port=443"
        result = extract_ips_from_text(text)
        assert "8.8.8.8" in result
        assert "1.1.1.1" in result

    def test_detect_csv_semicolon(self):
        from utils import detect_csv
        result = detect_csv("a;b;c\n1;2;3\n")
        assert result is not None
        delim, header, data = result
        assert delim == ";"
        assert header == "a;b;c"
        assert data == ["1;2;3"]

    def test_detect_csv_comma(self):
        from utils import detect_csv
        result = detect_csv("a,b\n1,2\n")
        assert result is not None
        assert result[0] == ","

    def test_detect_csv_not_csv(self):
        from utils import detect_csv
        assert detect_csv("just one line\n") is None
        assert detect_csv("no delimiters\nat all\n") is None

    def test_detect_ip_column_finds_ip_col(self):
        from utils import detect_ip_column
        rows = ["server1;8.8.8.8;443", "server2;1.1.1.1;80"]
        assert detect_ip_column(rows, ";") == 1

    def test_detect_ip_column_no_ip_returns_none(self):
        from utils import detect_ip_column
        rows = ["alice;hello", "bob;world"]
        assert detect_ip_column(rows, ";") is None

    def test_detect_ip_column_finds_ip_embedded_in_text(self):
        from utils import detect_ip_column
        rows = [
            "evt1;connection from 8.8.8.8 port 443",
            "evt2;request to 1.1.1.1 failed",
        ]
        assert detect_ip_column(rows, ";") == 1

    def test_parse_excel_non_xlsx_returns_none(self):
        from utils import parse_excel_to_csv
        assert parse_excel_to_csv(b"just plain text") is None
        assert parse_excel_to_csv(b"") is None

    def test_parse_excel_xlsx_roundtrip(self, tmp_path):
        import openpyxl
        from utils import parse_excel_to_csv

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["host", "ip", "port"])
        ws.append(["server1", "8.8.8.8", 443])
        ws.append(["server2", "1.1.1.1", 80])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        content = parse_excel_to_csv(path.read_bytes())
        assert content is not None
        lines = content.strip().splitlines()
        assert lines[0] == "host;ip;port"
        assert lines[1] == "server1;8.8.8.8;443"
        assert lines[2] == "server2;1.1.1.1;80"

    def test_parse_excel_sanitises_delimiter_in_cells(self, tmp_path):
        import openpyxl
        from utils import parse_excel_to_csv

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ip", "note"])
        ws.append(["8.8.8.8", "has;semicolon and\nnewline"])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        content = parse_excel_to_csv(path.read_bytes())
        assert content is not None
        lines = content.strip().splitlines()
        # semicolons and newlines inside cells are replaced with spaces,
        # so splitting on ';' still yields exactly 2 fields per row
        assert len(lines) == 2
        assert lines[0].split(";") == ["ip", "note"]
        assert lines[1].split(";")[0] == "8.8.8.8"
        note_field = lines[1].split(";")[1]
        assert "semicolon" in note_field
        assert "newline" in note_field
        assert "\n" not in note_field


# ---------------------------------------------------------------------------
# Provider unit tests (no network calls)
# ---------------------------------------------------------------------------

class TestBaseProviderKeyCache:
    """Test _load_api_key handling of directories (Docker bind mount artifact)."""

    def test_load_key_skips_directory(self, tmp_path):
        from providers.abuseipdb import AbuseIPDBProvider
        p = AbuseIPDBProvider()
        p.key_cache_file = ".enrichip-test-key"

        # Create a directory where the cache file would be
        dir_path = tmp_path / ".enrichip-test-key"
        dir_path.mkdir()

        with patch.object(type(p), '_get_key_path', return_value=dir_path):
            args = types.SimpleNamespace(use_abuseipdb=True, abuseipdb_api=None)
            result = p._load_api_key(args, 'abuseipdb_api', 'AbuseIPDB')

        # Should not crash, should return False (no key available)
        assert result is False

    def test_write_key_replaces_directory(self, tmp_path):
        from providers.abuseipdb import AbuseIPDBProvider
        p = AbuseIPDBProvider()
        p.key_cache_file = ".enrichip-test-key"

        # Create a directory where the cache file would be
        dir_path = tmp_path / ".enrichip-test-key"
        dir_path.mkdir()

        with patch.object(type(p), '_get_key_path', return_value=dir_path):
            args = types.SimpleNamespace(use_abuseipdb=True, abuseipdb_api="mykey123")
            result = p._load_api_key(args, 'abuseipdb_api', 'AbuseIPDB')

        assert result is True
        assert p.api_key == "mykey123"
        # Directory should have been replaced with a file
        assert dir_path.is_file()
        assert dir_path.read_text() == "mykey123"

    def test_load_key_from_file_works(self, tmp_path):
        from providers.abuseipdb import AbuseIPDBProvider
        p = AbuseIPDBProvider()

        key_file = tmp_path / ".enrichip-test-key"
        key_file.write_text("cached_key")

        with patch.object(type(p), '_get_key_path', return_value=key_file):
            args = types.SimpleNamespace(use_abuseipdb=True, abuseipdb_api=None)
            result = p._load_api_key(args, 'abuseipdb_api', 'AbuseIPDB')

        assert result is True
        assert p.api_key == "cached_key"


class TestBaseProviderProgressCallback:
    def test_progress_callback_defaults_to_none(self):
        from providers.dnsbl import DNSBLProvider
        p = DNSBLProvider()
        assert p.progress_callback is None


class TestDNSBLProvider:
    def test_headers(self):
        from providers.dnsbl import DNSBLProvider
        p = DNSBLProvider()
        assert p.get_headers() == ["DNSBL-Count"]

    def test_initialize_disabled(self):
        from providers.dnsbl import DNSBLProvider
        p = DNSBLProvider()
        args = types.SimpleNamespace(use_dnsbl=False)
        assert p.initialize(args) is True
        assert p.enabled is False

    def test_initialize_enabled(self):
        from providers.dnsbl import DNSBLProvider
        p = DNSBLProvider()
        args = types.SimpleNamespace(use_dnsbl=True)
        assert p.initialize(args) is True
        assert p.enabled is True

    def test_enrich_invalid_ip(self):
        from providers.dnsbl import DNSBLProvider
        p = DNSBLProvider()
        assert p.enrich("not-an-ip", {}) == [""]

    def test_enrich_private_ip(self):
        from providers.dnsbl import DNSBLProvider
        p = DNSBLProvider()
        assert p.enrich("192.168.1.1", {}) == [""]

    @patch("providers.dnsbl.socket.gethostbyname")
    def test_enrich_counts_hits(self, mock_dns):
        from providers.dnsbl import DNSBLProvider, DNSBL_SERVERS
        # Simulate 3 servers responding (hit) and the rest failing
        call_count = [0]
        def side_effect(query):
            call_count[0] += 1
            if call_count[0] <= 3:
                return "127.0.0.2"
            import socket as _socket
            raise _socket.gaierror("not found")
        mock_dns.side_effect = side_effect

        p = DNSBLProvider()
        result = p.enrich("8.8.8.8", {})
        assert result == ["3"]
        assert mock_dns.call_count == len(DNSBL_SERVERS)


class TestIPInfoProvider:
    def test_headers(self):
        from providers.ipinfo import IPInfoProvider
        p = IPInfoProvider()
        assert p.get_headers() == ["Hostname"]

    def test_enrich_invalid_ip(self):
        from providers.ipinfo import IPInfoProvider
        p = IPInfoProvider()
        assert p.enrich("garbage", {}) == [""]

    @patch("providers.ipinfo.urllib.request.urlopen")
    def test_enrich_stores_hostname_in_context(self, mock_urlopen):
        from providers.ipinfo import IPInfoProvider
        mock_resp = MagicMock()
        mock_resp.__enter__ = lambda s: s
        mock_resp.__exit__ = MagicMock(return_value=False)
        mock_resp.read.return_value = b'{"hostname": "dns.google"}'
        mock_urlopen.return_value = mock_resp

        p = IPInfoProvider()
        ctx = {}
        result = p.enrich("8.8.8.8", ctx)
        assert result == ["dns.google"]
        assert ctx["hostname"] == "dns.google"


class TestAbuseIPDBProvider:
    def test_headers(self):
        from providers.abuseipdb import AbuseIPDBProvider
        p = AbuseIPDBProvider()
        assert p.get_headers() == ["Abuse-Score", "Abuse-Reports", "Abuse-Country"]

    def test_enrich_invalid_ip(self):
        from providers.abuseipdb import AbuseIPDBProvider
        p = AbuseIPDBProvider()
        assert p.enrich("not-ip", {}) == ["", "", ""]


class TestProxycheckProvider:
    def test_headers(self):
        from providers.proxycheck import ProxycheckProvider
        p = ProxycheckProvider()
        assert len(p.get_headers()) == 9

    def test_enrich_invalid_ip(self):
        from providers.proxycheck import ProxycheckProvider
        p = ProxycheckProvider()
        assert p.enrich("nope", {}) == [""] * 9


class TestDNSDumpsterProvider:
    def test_headers(self):
        from providers.dnsdumpster import DNSDumpsterProvider
        p = DNSDumpsterProvider()
        assert p.get_headers() == ["Banners"]

    def test_enrich_no_hostname_in_context(self):
        from providers.dnsdumpster import DNSDumpsterProvider
        p = DNSDumpsterProvider()
        assert p.enrich("8.8.8.8", {}) == [""]
